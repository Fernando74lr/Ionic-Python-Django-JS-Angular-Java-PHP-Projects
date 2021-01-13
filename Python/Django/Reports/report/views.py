from .data.secret.credentials import server, username, password
from django.http.response import HttpResponse
from django.shortcuts import render
from .models import SqlServerConn
from openpyxl import Workbook
import datetime
import pyodbc

query_cache = []

def reports(request):
    # clients = get_clients(request)
    return render(request, 'report/reports.html')#, {'SqlServerConn' : clients})

def get_clients(request):
    enterprises = [
        'adCOM_GRUPO_SANTA_MARIA_SA',
        'adCOM_KL_13_SERVICIOS_SA',
        'adCOM_YUNESGO_ASOCIADOS_S'
    ]

    global query_cache
    result = []

    if len(query_cache) == 0:
        print("\nGETTING DATA\n")

        for database in enterprises:
            conn = pyodbc.connect('DRIVER={sql server};'+
                                'SERVER=' + server + ';'+
                                'DATABASE=' + database + ';'+
                                'UID=' + username + ';'+
                                'PWD=' + password + ';')
            cursor = conn.cursor()
            cursor.execute("SELECT "+
                        "CIDCLIENTEPROVEEDOR, CRAZONSOCIAL, CRFC,"+
                        "CFECHAALTA, CESTATUS, CIDAGENTEVENTA "+
                        "FROM admClientes")
            temp = []
            for query in cursor.fetchall():            
                query = list(query)
                query.insert(0, database.replace('adCOM_', '').replace('_', ' '))
                query.insert(0, database)
                temp.append(query)
            result.append(temp)
        
        query_cache = result
    else:
        print("\nUSING CACHE\n")
        result = query_cache
    
    if len(result) == 3:
        print("\nALL RIGHT\n")
    else:
        print("SOMETHING WENT WRONG")

    # return result
    return render(request, 'report/reports.html', {'SqlServerConn' : result})

def clients_report(request):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'REPORTE CLIENTES'
    ws.merge_cells('A1:B1')
    ws['C1'] = datetime.datetime.now()

    # DATA
    ws['A3'] = 'ID CLIENTE'
    ws['B3'] = 'CORPORACIÓN'
    ws['C3'] = 'RAZÓN SOCIAL'
    ws['D3'] = 'RFC'
    ws['E3'] = 'ID AGENTE VENTA'
    ws['F3'] = 'FECHA ALTA'
    ws['G3'] = 'ESTATUS'
    ws['H3'] = 'BASE DE DATOS'

    global query_cache
    counter = 4

    if len(query_cache) > 0:
        print("\nCREATING EXCEL\n")
        for enterprise in query_cache:
            for data in enterprise:
                # ID Cliente
                ws.cell(row=counter,column=1).value = data[2]
                # Corporación
                ws.cell(row=counter,column=2).value = data[1]
                # Razón Social
                ws.cell(row=counter,column=3).value = data[3]
                # RFC
                ws.cell(row=counter,column=4).value = data[4]
                # ID Agente Venta
                ws.cell(row=counter,column=5).value = data[7]
                # Fecha Alta
                ws.cell(row=counter,column=6).value = str(data[5]).split(' ')[0]
                # Estatus
                ws.cell(row=counter,column=7).value = data[6]
                # Base de datos
                ws.cell(row=counter,column=8).value = data[0]

                counter+=1
    else:
        print("\nCACHE EMPTY\n")

    filename = 'Reporte_Clientes.xlsx'
    response = HttpResponse(content_type = 'application/ms-excel')
    content = 'attachment; filename = {0}'.format(filename)
    response['Content-Disposition'] = content
    wb.save(response)

    return response